import pandas as pd
import numpy as np
from datetime import datetime, timedelta, date
from typing import Dict, List, Tuple, Optional
import sys
import os
import logging
from sqlalchemy import func, and_, or_

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.db_utils import (
    SessionLocal, Student, AttendanceSession, AttendanceRecord, 
    Class, Subject, get_active_academic_year, record_performance_metric
)

logger = logging.getLogger(__name__)

class AttendanceAnalytics:
    """Comprehensive attendance analytics and reporting."""
    
    def __init__(self):
        self.session = SessionLocal()
    
    def close(self):
        self.session.close()
    
    def calculate_student_attendance_percentage(self, student_id: int, 
                                              class_id: int = None, 
                                              subject_id: int = None,
                                              start_date: date = None, 
                                              end_date: date = None,
                                              days_back: int = None) -> float:
        """Calculate attendance percentage for a student."""
        try:
            query = self.session.query(AttendanceRecord).join(AttendanceSession).filter(
                AttendanceRecord.student_id == student_id
            )
            
            # Apply filters
            if class_id:
                query = query.filter(AttendanceSession.class_id == class_id)
            
            if subject_id:
                query = query.filter(AttendanceSession.subject_id == subject_id)
            
            if days_back:
                start_date = datetime.now().date() - timedelta(days=days_back)
                query = query.filter(AttendanceSession.date >= start_date)
            
            if start_date:
                query = query.filter(AttendanceSession.date >= start_date)
            
            if end_date:
                query = query.filter(AttendanceSession.date <= end_date)
            
            total_records = query.count()
            present_records = query.filter(AttendanceRecord.status == 'present').count()
            
            if total_records == 0:
                return 0.0
            
            return (present_records / total_records) * 100
            
        except Exception as e:
            logger.error(f"Error calculating attendance percentage: {e}")
            return 0.0
    
    def get_class_attendance_summary(self, class_id: int, 
                                   start_date: date = None, 
                                   end_date: date = None) -> Dict:
        """Get attendance summary for a class."""
        try:
            # Default to current month if no dates provided
            if not start_date:
                start_date = datetime.now().replace(day=1).date()
            if not end_date:
                end_date = datetime.now().date()
            
            students = self.session.query(Student).filter(
                Student.class_id == class_id,
                Student.is_active == True
            ).all()
            
            summary = {
                'total_students': len(students),
                'period': f"{start_date} to {end_date}",
                'students': []
            }
            
            for student in students:
                attendance_percentage = self.calculate_student_attendance_percentage(
                    student.id, class_id, start_date=start_date, end_date=end_date
                )
                
                # Get total sessions and present count
                total_query = self.session.query(AttendanceRecord).join(AttendanceSession).filter(
                    AttendanceRecord.student_id == student.id,
                    AttendanceSession.class_id == class_id,
                    AttendanceSession.date >= start_date,
                    AttendanceSession.date <= end_date
                )
                
                total_sessions = total_query.count()
                present_sessions = total_query.filter(AttendanceRecord.status == 'present').count()
                absent_sessions = total_query.filter(AttendanceRecord.status == 'absent').count()
                late_sessions = total_query.filter(AttendanceRecord.status == 'late').count()
                
                summary['students'].append({
                    'id': student.id,
                    'name': student.name,
                    'roll': student.roll,
                    'total_sessions': total_sessions,
                    'present': present_sessions,
                    'absent': absent_sessions,
                    'late': late_sessions,
                    'percentage': attendance_percentage,
                    'status': self._get_attendance_status(attendance_percentage)
                })
            
            # Calculate class averages
            if summary['students']:
                percentages = [s['percentage'] for s in summary['students']]
                summary['class_average'] = np.mean(percentages)
                summary['highest_attendance'] = max(percentages)
                summary['lowest_attendance'] = min(percentages)
                summary['students_below_75'] = len([p for p in percentages if p < 75])
            
            return summary
            
        except Exception as e:
            logger.error(f"Error getting class attendance summary: {e}")
            return {}
    
    def get_attendance_trends(self, class_id: int = None, 
                            student_id: int = None, 
                            days_back: int = 30) -> Dict:
        """Get attendance trends over time."""
        try:
            end_date = datetime.now().date()
            start_date = end_date - timedelta(days=days_back)
            
            # Use a simpler approach without problematic CASE statements
            base_query = self.session.query(AttendanceSession).filter(
                AttendanceSession.date >= start_date,
                AttendanceSession.date <= end_date
            )
            
            if class_id:
                base_query = base_query.filter(AttendanceSession.class_id == class_id)
            
            sessions = base_query.all()
            
            # Group by date
            date_stats = {}
            
            for session in sessions:
                date_str = session.date.strftime('%Y-%m-%d')
                if date_str not in date_stats:
                    date_stats[date_str] = {
                        'total': 0,
                        'present': 0,
                        'absent': 0,
                        'late': 0
                    }
                
                # Get records for this session
                records_query = self.session.query(AttendanceRecord).filter(
                    AttendanceRecord.session_id == session.id
                )
                
                if student_id:
                    records_query = records_query.filter(AttendanceRecord.student_id == student_id)
                
                records = records_query.all()
                
                for record in records:
                    date_stats[date_str]['total'] += 1
                    if record.status == 'present':
                        date_stats[date_str]['present'] += 1
                    elif record.status == 'absent':
                        date_stats[date_str]['absent'] += 1
                    elif record.status == 'late':
                        date_stats[date_str]['late'] += 1
            
            # Convert to lists
            trends = {
                'dates': [],
                'total': [],
                'present': [],
                'absent': [],
                'late': [],
                'percentages': []
            }
            
            # Sort dates
            sorted_dates = sorted(date_stats.keys())
            
            for date_str in sorted_dates:
                stats = date_stats[date_str]
                trends['dates'].append(date_str)
                trends['total'].append(stats['total'])
                trends['present'].append(stats['present'])
                trends['absent'].append(stats['absent'])
                trends['late'].append(stats['late'])
                
                if stats['total'] > 0:
                    percentage = float(stats['present'] / stats['total'] * 100)
                else:
                    percentage = 0.0
                trends['percentages'].append(percentage)
            
            # Ensure we always return valid structure even if no data
            if not trends['dates']:
                # Generate empty data for the date range
                current = start_date
                while current <= end_date:
                    trends['dates'].append(current.strftime('%Y-%m-%d'))
                    trends['total'].append(0)
                    trends['present'].append(0)
                    trends['absent'].append(0)
                    trends['late'].append(0)
                    trends['percentages'].append(0.0)
                    current += timedelta(days=1)
            
            return trends
            
        except Exception as e:
            logger.error(f"Error getting attendance trends: {e}")
            # Return empty but valid structure
            return {
                'dates': [],
                'total': [],
                'present': [],
                'absent': [],
                'late': [],
                'percentages': []
            }
    
    def get_subject_wise_attendance(self, student_id: int) -> Dict:
        """Get subject-wise attendance for a student."""
        try:
            subjects_query = self.session.query(Subject).join(Class).join(Student).filter(
                Student.id == student_id,
                Subject.is_active == True
            )
            
            subject_attendance = {}
            
            for subject in subjects_query:
                attendance_percentage = self.calculate_student_attendance_percentage(
                    student_id, subject_id=subject.id
                )
                
                # Get session details
                total_sessions = self.session.query(AttendanceRecord).join(AttendanceSession).filter(
                    AttendanceRecord.student_id == student_id,
                    AttendanceSession.subject_id == subject.id
                ).count()
                
                present_sessions = self.session.query(AttendanceRecord).join(AttendanceSession).filter(
                    AttendanceRecord.student_id == student_id,
                    AttendanceSession.subject_id == subject.id,
                    AttendanceRecord.status == 'present'
                ).count()
                
                subject_attendance[subject.name] = {
                    'subject_id': subject.id,
                    'subject_code': subject.code,
                    'total_sessions': total_sessions,
                    'present_sessions': present_sessions,
                    'percentage': attendance_percentage,
                    'status': self._get_attendance_status(attendance_percentage)
                }
            
            return subject_attendance
            
        except Exception as e:
            logger.error(f"Error getting subject-wise attendance: {e}")
            return {}
    
    def get_low_attendance_students(self, threshold: float = 75, 
                                  class_id: int = None, 
                                  days_back: int = 30) -> List[Dict]:
        """Get students with attendance below threshold."""
        try:
            query = self.session.query(Student).filter(Student.is_active == True)
            
            if class_id:
                query = query.filter(Student.class_id == class_id)
            
            students = query.all()
            low_attendance_students = []
            
            for student in students:
                attendance_percentage = self.calculate_student_attendance_percentage(
                    student.id, days_back=days_back
                )
                
                if attendance_percentage < threshold:
                    low_attendance_students.append({
                        'id': student.id,
                        'name': student.name,
                        'roll': student.roll,
                        'class_id': student.class_id,
                        'percentage': attendance_percentage,
                        'email': student.email,
                        'parent_email': student.parent_email
                    })
            
            # Sort by lowest attendance first
            low_attendance_students.sort(key=lambda x: x['percentage'])
            
            return low_attendance_students
            
        except Exception as e:
            logger.error(f"Error getting low attendance students: {e}")
            return []
    
    def get_daily_attendance_summary(self, class_id: int, target_date: date) -> Dict:
        """Get attendance summary for a specific day."""
        try:
            sessions = self.session.query(AttendanceSession).filter(
                AttendanceSession.class_id == class_id,
                func.date(AttendanceSession.date) == target_date
            ).all()
            
            if not sessions:
                return {}
            
            summary = {}
            
            for session in sessions:
                records = self.session.query(AttendanceRecord).filter(
                    AttendanceRecord.session_id == session.id
                ).all()
                
                session_summary = {
                    'session_name': session.session_name,
                    'subject': session.subject.name if session.subject else 'General',
                    'total_students': len(records),
                    'present': len([r for r in records if r.status == 'present']),
                    'absent': len([r for r in records if r.status == 'absent']),
                    'late': len([r for r in records if r.status == 'late']),
                    'students': []
                }
                
                for record in records:
                    session_summary['students'].append({
                        'name': record.student.name,
                        'roll': record.student.roll,
                        'status': record.status,
                        'confidence': record.confidence,
                        'marked_at': record.marked_at.strftime('%H:%M:%S') if record.marked_at else 'N/A'
                    })
                
                if session_summary['total_students'] > 0:
                    session_summary['attendance_rate'] = (
                        session_summary['present'] / session_summary['total_students']
                    ) * 100
                
                summary[session.id] = session_summary
            
            return summary
            
        except Exception as e:
            logger.error(f"Error getting daily attendance summary: {e}")
            return {}
    
    def generate_attendance_report(self, class_id: int = None, 
                                 student_id: int = None,
                                 start_date: date = None, 
                                 end_date: date = None,
                                 report_type: str = 'summary') -> Dict:
        """Generate comprehensive attendance report."""
        try:
            if not start_date:
                start_date = datetime.now().replace(day=1).date()
            if not end_date:
                end_date = datetime.now().date()
            
            report = {
                'report_type': report_type,
                'period': f"{start_date} to {end_date}",
                'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'filters': {
                    'class_id': class_id,
                    'student_id': student_id
                }
            }
            
            if report_type == 'summary' and class_id:
                report['data'] = self.get_class_attendance_summary(class_id, start_date, end_date)
                
            elif report_type == 'student' and student_id:
                student = self.session.query(Student).filter(Student.id == student_id).first()
                if student:
                    report['data'] = {
                        'student_info': {
                            'name': student.name,
                            'roll': student.roll,
                            'class': student.class_obj.name if student.class_obj else 'N/A'
                        },
                        'overall_percentage': self.calculate_student_attendance_percentage(
                            student_id, start_date=start_date, end_date=end_date
                        ),
                        'subject_wise': self.get_subject_wise_attendance(student_id),
                        'trends': self.get_attendance_trends(student_id=student_id)
                    }
                    
            elif report_type == 'trends':
                report['data'] = self.get_attendance_trends(
                    class_id=class_id, 
                    student_id=student_id,
                    days_back=(end_date - start_date).days
                )
                
            elif report_type == 'low_attendance':
                report['data'] = self.get_low_attendance_students(
                    class_id=class_id,
                    days_back=(end_date - start_date).days
                )
            
            return report
            
        except Exception as e:
            logger.error(f"Error generating attendance report: {e}")
            return {}
    
    def get_attendance_statistics(self, class_id: int = None) -> Dict:
        """Get comprehensive attendance statistics."""
        try:
            stats = {}
            
            # Overall statistics
            if class_id:
                query_filter = {'class_id': class_id}
            else:
                query_filter = {}
            
            # Total students
            total_students = self.session.query(Student).filter(
                Student.is_active == True,
                *[getattr(Student, k) == v for k, v in query_filter.items()]
            ).count()
            
            # Total sessions
            total_sessions = self.session.query(AttendanceSession).filter(
                *[getattr(AttendanceSession, k) == v for k, v in query_filter.items()]
            ).count()
            
            # Attendance rates
            attendance_query = self.session.query(AttendanceRecord).join(AttendanceSession)
            if class_id:
                attendance_query = attendance_query.filter(AttendanceSession.class_id == class_id)
            
            total_records = attendance_query.count()
            present_records = attendance_query.filter(AttendanceRecord.status == 'present').count()
            
            stats = {
                'total_students': total_students,
                'total_sessions': total_sessions,
                'total_records': total_records,
                'overall_attendance_rate': (present_records / total_records * 100) if total_records > 0 else 0,
                'students_above_90': 0,
                'students_75_90': 0,
                'students_below_75': 0
            }
            
            # Categorize students by attendance
            if class_id:
                students = self.session.query(Student).filter(
                    Student.class_id == class_id,
                    Student.is_active == True
                ).all()
            else:
                students = self.session.query(Student).filter(Student.is_active == True).all()
            
            for student in students:
                percentage = self.calculate_student_attendance_percentage(student.id, class_id)
                if percentage >= 90:
                    stats['students_above_90'] += 1
                elif percentage >= 75:
                    stats['students_75_90'] += 1
                else:
                    stats['students_below_75'] += 1
            
            return stats
            
        except Exception as e:
            logger.error(f"Error getting attendance statistics: {e}")
            return {}
    
    def _get_attendance_status(self, percentage: float) -> str:
        """Get attendance status based on percentage."""
        if percentage >= 90:
            return 'Excellent'
        elif percentage >= 75:
            return 'Good'
        elif percentage >= 60:
            return 'Average'
        else:
            return 'Poor'

def calculate_student_attendance_percentage(session, student_id: int, days_back: int = 30) -> float:
    """Standalone function to calculate student attendance percentage."""
    analytics = AttendanceAnalytics()
    try:
        return analytics.calculate_student_attendance_percentage(
            student_id, days_back=days_back
        )
    finally:
        analytics.close()

def get_daily_attendance_summary(session, class_id: int, target_date: date) -> Dict:
    """Standalone function to get daily attendance summary."""
    analytics = AttendanceAnalytics()
    try:
        return analytics.get_daily_attendance_summary(class_id, target_date)
    finally:
        analytics.close()

def export_attendance_data_to_excel(data: Dict, filename: str) -> str:
    """Export attendance data to Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Headers
        headers = ['Student Name', 'Roll Number', 'Total Sessions', 'Present', 'Absent', 'Late', 'Percentage', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        if 'students' in data:
            for row, student in enumerate(data['students'], 2):
                ws.cell(row=row, column=1, value=student['name'])
                ws.cell(row=row, column=2, value=student['roll'])
                ws.cell(row=row, column=3, value=student['total_sessions'])
                ws.cell(row=row, column=4, value=student['present'])
                ws.cell(row=row, column=5, value=student['absent'])
                ws.cell(row=row, column=6, value=student['late'])
                ws.cell(row=row, column=7, value=f"{student['percentage']:.1f}%")
                ws.cell(row=row, column=8, value=student['status'])
                
                # Color coding based on attendance
                if student['percentage'] >= 90:
                    fill_color = 'C6EFCE'  # Light green
                elif student['percentage'] >= 75:
                    fill_color = 'FFEB9C'  # Light yellow
                else:
                    fill_color = 'FFC7CE'  # Light red
                
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color=fill_color, 
                        end_color=fill_color, 
                        fill_type='solid'
                    )
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        os.makedirs("data/exports", exist_ok=True)
        filepath = os.path.join("data/exports", filename)
        wb.save(filepath)
        
        return filepath
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        return None

def monitor_system_performance():
    """Monitor and record system performance metrics."""
    try:
        import psutil
        
        session = SessionLocal()
        
        # CPU usage
        cpu_percent = psutil.cpu_percent(interval=1)
        record_performance_metric(session, 'cpu_usage', cpu_percent, 'percent', 'system')
        
        # Memory usage
        memory = psutil.virtual_memory()
        record_performance_metric(session, 'memory_usage', memory.percent, 'percent', 'system')
        record_performance_metric(session, 'memory_available', memory.available / (1024**3), 'GB', 'system')
        
        # Disk usage
        disk = psutil.disk_usage('/')
        disk_percent = (disk.used / disk.total) * 100
        record_performance_metric(session, 'disk_usage', disk_percent, 'percent', 'system')
        
        logger.info(f"Performance metrics recorded: CPU {cpu_percent}%, Memory {memory.percent}%, Disk {disk_percent}%")
        
        session.close()
        
    except Exception as e:
        logger.error(f"Error monitoring system performance: {e}") 